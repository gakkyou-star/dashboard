\
# -*- coding: utf-8 -*-
"""
鹿追高校ダッシュボード 集計スクリプト

マスターデータフォルダ内の元データ（Excel/CSV）を集計し、
表示用の output/dashboard_data.json を生成する。

- カテゴリごとに関数を分割。データ追加・修正時は該当関数のみ差分編集すればよい。
- 特定ファイル名に強く依存しないよう、glob パターンでファイルを探索し、
  複数候補がある場合は更新日時が最も新しいものを採用する。
- 個人名・住所などの個人情報は集計後の件数・割合のみを出力し、
  生データ（氏名・住所そのもの）は json に含めない。
"""

import glob
import json
import os
import re
import unicodedata
import warnings
from datetime import datetime

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", ".."))  # マスターデータ フォルダ
OUTPUT_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "output"))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "dashboard_data.json")
OUTPUT_JS_FILE = os.path.join(OUTPUT_DIR, "dashboard_data.js")

SCHOOL_NAME = "鹿追高校"


def latest_file(patterns):
    """glob パターン（複数可）にマッチするファイルのうち、更新日時が最新のものを返す。"""
    candidates = []
    for pattern in patterns:
        candidates.extend(glob.glob(os.path.join(ROOT_DIR, pattern), recursive=True))
    candidates = [c for c in candidates if os.path.isfile(c) and not os.path.basename(c).startswith("~$")]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def not_ready(note="データ準備中"):
    return {"ready": False, "note": note}


# ---------------------------------------------------------------------------
# 1. 学校概況
# ---------------------------------------------------------------------------
def aggregate_school_overview():
    result = {}

    # --- 出身地域別内訳（町内／十勝管内／道内／道外）・入学者推移（経年） -------
    # 「鹿高入学者推移.xlsx」の「入学者内訳」シートは2022〜2025年度分のみで
    # 2026年度が欠けているため、全年度（1986〜）を持つ「データ」シートの
    # 町内入学者／管内入学者／道内入学者／道外入学者列から一括で導出する。
    # （道内・道外入学者は2023年度の地域みらい留学制度加盟以降のみ値が入る）
    # TODO: 生徒ごとの住所データが揃ったら、infer_prefecture() を使って
    #       住所ベースの出身地別内訳に切り替える（現時点ではファイル未提供のため未実装）。
    trend_path = latest_file(["高校/鹿高入学者推移*.xlsx"])
    origin_years = []
    series = []
    origin_categories = [("町内", "町内入学者"), ("管内", "管内入学者"), ("道内", "道内入学者"), ("道外", "道外入学者")]
    if trend_path:
        try:
            df = pd.read_excel(trend_path, sheet_name="データ")
            df = df[pd.to_numeric(df["西暦"], errors="coerce").notna()]
            df["西暦"] = df["西暦"].astype(int)
            df = df[(df["西暦"] >= 1986) & (df["西暦"] <= 2100)]
            df = df.sort_values("西暦")
            for _, row in df.iterrows():
                if pd.isna(row.get("鹿高入学者人数")):
                    continue
                year = int(row["西暦"])
                series.append({
                    "year": year,
                    "count": int(row["鹿高入学者人数"]),
                    "local_rate": None if pd.isna(row.get("地元中学校からの入学率")) else round(float(row["地元中学校からの入学率"]) * 100, 1),
                })
                breakdown = [
                    {"label": label, "count": int(row[col])}
                    for label, col in origin_categories
                    if col in row and not pd.isna(row[col])
                ]
                if breakdown:
                    origin_years.append({"year": year, "total": sum(b["count"] for b in breakdown), "breakdown": breakdown})
            result["admission_trend"] = {"ready": True, "series": series}
        except Exception:
            result["admission_trend"] = not_ready()
    else:
        result["admission_trend"] = not_ready()

    # --- 在籍数・学年構成（推計） -------------------------------------------
    # 生徒一覧.xlsx（生徒/生徒一覧.xlsx）は2023-09-14時点の実数だが更新が止まっており
    # 現況を反映していないため、より新しい「鹿高入学者推移.xlsx」の入学者数（直近3年度）を
    # 各学年の在籍数として代用する推計値を採用する（転入出等の差は反映されない）。
    recent_series = series[-3:]
    if recent_series:
        grade_labels = ["1年", "2年", "3年"]
        by_grade = [
            {"label": label, "count": recent_series[-(i + 1)]["count"]}
            for i, label in enumerate(grade_labels)
            if i < len(recent_series)
        ]
        result["enrollment"] = {
            "ready": True,
            "total": sum(g["count"] for g in by_grade),
            "by_grade": by_grade,
            "estimated": True,
            "source_updated": datetime.fromtimestamp(os.path.getmtime(trend_path)).strftime("%Y-%m-%d"),
        }
    else:
        result["enrollment"] = not_ready()

    if origin_years:
        result["origin_region"] = {"ready": True, "years": origin_years}
    else:
        result["origin_region"] = not_ready()

    # --- 在籍生徒の町内割合（現1〜3年生=直近3年度の入学者を合算） -------------
    recent3 = origin_years[-3:]
    if recent3:
        totals = {}
        for y in recent3:
            for b in y["breakdown"]:
                totals[b["label"]] = totals.get(b["label"], 0) + b["count"]
        result["current_enrollment_origin"] = {
            "ready": True,
            "years": [y["year"] for y in recent3],
            "breakdown": [{"label": label, "count": totals.get(label, 0)} for label, _ in origin_categories],
        }
    else:
        result["current_enrollment_origin"] = not_ready()

    # --- 入学理由アンケート結果 ---------------------------------------------
    reason_path = latest_file(["高校/入学理由/*入学理由*アンケート*.xlsx"])
    if reason_path:
        try:
            df = pd.read_excel(reason_path, sheet_name=0)
            n = len(df)
            col1 = "鹿追高校に出願した「１番の理由」は何ですか？"
            col_support = "鹿追町からの支援で最も魅力的なものはどれですか？"

            def tally(col):
                if col not in df.columns:
                    return []
                counts = df[col].dropna().value_counts()
                return [{"label": k, "count": int(v)} for k, v in counts.items()]

            result["admission_reason"] = {
                "ready": True,
                "respondents": int(n),
                "top_reason": tally(col1),
                "attractive_support": tally(col_support),
                "source_updated": datetime.fromtimestamp(os.path.getmtime(reason_path)).strftime("%Y-%m-%d"),
            }
        except Exception:
            result["admission_reason"] = not_ready()
    else:
        result["admission_reason"] = not_ready()

    return result


# ---------------------------------------------------------------------------
# 2. 教育活動
# ---------------------------------------------------------------------------
def aggregate_education_activities(charm_scores):
    result = {}

    # 探究学習の成果は「高校魅力化評価」アンケートの探究性スコアを代理指標として利用
    if charm_scores.get("ready"):
        inquiry = [c for c in charm_scores["categories"] if "探究性" in c["label"]]
        result["inquiry_learning"] = {
            "ready": True,
            "note": "探究性に関わる学習活動・学習環境・自己認識・行動・ウェルビーイングの各評価スコア（生徒アンケート）",
            "scores": inquiry,
        }
    else:
        result["inquiry_learning"] = not_ready()

    result["international_exchange"] = not_ready("件数・成果データは今後追加予定")
    result["secondary_collaboration"] = not_ready("件数・成果データは今後追加予定")
    result["community_collaboration"] = not_ready("件数・成果データは今後追加予定")

    return result


# ---------------------------------------------------------------------------
# 3. 進路・学習
# ---------------------------------------------------------------------------
# 進路先カテゴリ別推移グラフの系列順（ドーナツグラフと同じ配色に対応させるため固定順とする）
CAREER_CATEGORY_ORDER = ["専門学校", "就職", "私立大", "国公立", "その他", "短大"]
# 上記6区分に含まれない進路（年度により粒度が異なる区分）は「その他」に統合する
CAREER_CATEGORY_ALIASES = {"高看": "その他", "留学": "その他"}


def aggregate_career_and_learning():
    result = {}

    # --- 進路決定率・進路実績 ------------------------------------------------
    career_path = latest_file(["高校/進路実績*.xlsx"])
    if career_path:
        try:
            df = pd.read_excel(career_path, sheet_name="テーブル2").dropna(how="all")
            df = df[pd.to_numeric(df["年度"], errors="coerce").notna()]
            df["年度"] = df["年度"].astype(int)
            latest_year = int(df["年度"].max())
            latest = df[df["年度"] == latest_year]
            latest_total = int(latest["人数"].sum())
            latest_breakdown = (
                latest.groupby("進路")["人数"].sum().sort_values(ascending=False)
            )

            # 進路先カテゴリ別・年度別の推移（現行の6区分に正規化。「高看」「留学」は「その他」に統合）
            df["進路正規化"] = df["進路"].map(lambda x: CAREER_CATEGORY_ALIASES.get(x, x))
            pivot = df.groupby(["年度", "進路正規化"])["人数"].sum().unstack(fill_value=0)
            years_list = sorted(int(y) for y in pivot.index)
            category_trend = {
                "years": years_list,
                "categories": [
                    {
                        "label": cat,
                        "data": [int(pivot.loc[y, cat]) if cat in pivot.columns else 0 for y in years_list],
                    }
                    for cat in CAREER_CATEGORY_ORDER
                ],
            }

            result["career_outcome"] = {
                "ready": True,
                "latest_year": latest_year,
                "decision_rate": 100.0 if latest_total > 0 else None,
                "total": latest_total,
                "breakdown": [
                    {"label": k, "count": int(v)} for k, v in latest_breakdown.items() if v > 0
                ],
                "category_trend": category_trend,
            }
        except Exception:
            result["career_outcome"] = not_ready()
    else:
        result["career_outcome"] = not_ready()

    # --- 外部検定（英検）結果集計 --------------------------------------------
    eiken_path = latest_file(["学力/英検/*英検*.xlsx"])
    if eiken_path:
        try:
            df = pd.read_excel(eiken_path, sheet_name="英検データ")
            hs = df[df["学校名"] == SCHOOL_NAME].copy()
            total_students = int(len(hs))
            holders = hs[hs["保有級"].notna()]
            order = ["1", "1（準）", "2", "2（準）", "3", "4", "5"]
            counts = holders["保有級"].value_counts()
            grade_counts = [{"label": g, "count": int(counts.get(g, 0))} for g in order if counts.get(g, 0) > 0]
            result["eiken"] = {
                "ready": True,
                "total_students": total_students,
                "holders": int(len(holders)),
                "grade_breakdown": grade_counts,
                "source_updated": datetime.fromtimestamp(os.path.getmtime(eiken_path)).strftime("%Y-%m-%d"),
            }
        except Exception:
            result["eiken"] = not_ready()
    else:
        result["eiken"] = not_ready()

    result["mock_exam"] = not_ready("模試・学習状況データは今後追加予定")

    # --- 魅力化アンケート（探究学習等の評価結果） ----------------------------
    charm_path = latest_file(["高校/高校魅力化評価/*.xlsx"])
    charm_result = not_ready()
    if charm_path:
        try:
            raw = pd.read_excel(charm_path, sheet_name="Worksheet", header=None)
            categories = []
            years_header = None
            for i, row in raw.iterrows():
                label = row[0]
                if isinstance(label, str) and label.startswith("回答年度"):
                    pass
                if isinstance(label, str) and label.startswith("【"):
                    def pct(v):
                        if isinstance(v, str) and v.endswith("%"):
                            return float(v[:-1])
                        return None
                    categories.append({
                        "label": label.strip("【】"),
                        "school_latest": pct(row[4]),
                        "school_prev": pct(row[3]),
                        "national_latest": pct(row[12]),
                    })
            respondents_row = raw[raw[0] == "質問事項"]
            respondents = None
            if not respondents_row.empty:
                val = respondents_row.iloc[0][4]
                if isinstance(val, str) and val.endswith("人"):
                    respondents = int(val[:-1])
            if categories:
                traits = ["主体性", "協働性", "探究性", "社会性"]
                trait_averages = []
                for trait in traits:
                    matched = [c for c in categories if c["label"].startswith(trait)]
                    matched = [c for c in matched if c["school_latest"] is not None]
                    if matched:
                        trait_averages.append({
                            "label": trait,
                            "school_latest": round(sum(c["school_latest"] for c in matched) / len(matched), 1),
                            "national_latest": round(sum(c["national_latest"] for c in matched if c["national_latest"] is not None) / len(matched), 1),
                        })
                charm_result = {
                    "ready": True,
                    "respondents": respondents,
                    "trait_averages": trait_averages,
                    "categories": categories,
                    "source_updated": datetime.fromtimestamp(os.path.getmtime(charm_path)).strftime("%Y-%m-%d"),
                }
        except Exception:
            pass
    result["charm_survey"] = charm_result

    return result, charm_result


# ---------------------------------------------------------------------------
# 4. 財政支援
# ---------------------------------------------------------------------------
# 財源内訳の列位置（「R8_前年比」シートの列番号。1始まり。国/道/地方債/その他/特別交付税は
# 交付金等として合算し、一財（一般財源）とのみ区別する）
FUND_COLUMNS = {"国": 13, "道": 14, "地方債": 15, "その他": 16, "特別交付税": 17, "一財": 18}

# 支える会総会資料（管理文書/38_学校教育その他/高校関係/高校を支える会/2026総会資料/
# 支える会資料_令和8年度.docx）の「高校支援の内容」表にある「支援項目」区分。
# 財政支援セクションは元データの区分・小計単位ではなく、この支援項目で整理する。
SUPPORT_CATEGORY_ORDER = ["学習支援", "部活動等支援", "通学環境整備", "経済支援", "進路支援", "入学者対策"]

# 「R8_前年比」シートの「内容」列（1行=1予算科目）を、上記の支援項目・事業ラベルに
# 対応させるための行単位マッピング。元データの「小計」バンドルは支援項目をまたぐ
# ことがある（例:「協力会補助事業」に全国募集の広報費と公設塾運営費が混在、
# 「通学費外助成事業」に通学費補助と見学旅行費が混在）ため、小計単位ではなく
# 内容行単位で分類してから事業ラベル単位で再集計する。
# 内容の文言は年度更新で変わり得るため、次年度ファイルで一致しなくなった場合は
# 「未分類」区分に集計される（ログにも出力する）ので、その件数を見て本マッピングを
# 更新すること。
FINANCE_CONTENT_CATEGORY = {
    "職員カナダ引率": ("学習支援", "カナダ短期留学（海外派遣事業）"),
    "海外派遣事業": ("学習支援", "カナダ短期留学（海外派遣事業）"),
    "オンライン公設塾英語検定受験料助成": ("学習支援", "オンライン公設塾"),
    "オンライン公設塾委託料": ("学習支援", "オンライン公設塾"),
    "協力会補助金（公設塾運営費）": ("学習支援", "オンライン公設塾"),
    "高校インターネット料": ("学習支援", "高校インターネット料"),
    "ストニィプレイン町派遣事業負担金": ("学習支援", "カナダ短期留学（海外派遣事業）"),
    "ストニィプレイン町受入事業負担金": ("学習支援", "カナダ短期留学（海外派遣事業）"),
    "生徒用タブレット整備事業": ("学習支援", "生徒用タブレット整備事業"),

    "高校教育振興補助金": ("部活動等支援", "高校教育振興補助金"),

    "協力会補助金（通学車両運行）": ("通学環境整備", "通学バス運行"),
    "協力会補助金（部活バス運行）": ("通学環境整備", "通学バス運行"),
    "通学費外助成（通学費）": ("通学環境整備", "通学費外助成（通学費・下宿）"),
    "通学費外助成（下宿）": ("通学環境整備", "通学費外助成（通学費・下宿）"),
    "コーディネーター（宗雲③）": ("入学者対策", "コーディネーター任用"),
    "鹿追高等学校女子専用下宿管理委託料": ("通学環境整備", "女子寮運営（委託）"),
    "シェアハウス調理員報酬": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス調理員手当": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス通勤手当": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス消耗品費": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス燃料費（3棟）": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス水道代（5戸）": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス修繕料": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス賄材料費(生徒33名・ﾊｳｽﾏｽﾀｰ2名)": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "インターネット使用料（11ヵ所）": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "調理員便検査": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス管理委託料（和田）": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス食事棟改修実施設計": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウス外構緑化工事": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "シェアハウスマスター（和田③、播磨②）": ("通学環境整備", "シェアハウス運営（Penguin House）"),
    "キュービクル管理委託料": ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）"),
    "AED使用料": ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）"),
    "緑化外構工事": ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）"),
    "寄宿舎燃料費（居住棟4棟、食堂1棟）": ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）"),
    "寄宿舎水道代（5戸）": ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）"),
    "管理運営業務委託料": ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）"),

    "通学費外助成（見学旅行）": ("経済支援", "入学準備・見学旅行助成"),
    "武藤孔二記念奨学補助金": ("経済支援", "武藤孔二記念奨学補助金"),

    "修学資金貸付金": ("進路支援", "修学資金貸付金"),

    "協力会補助金（全国募集事業費等(広報))": ("入学者対策", "協力会補助金（全国募集事業費等）"),
    "協力会補助金（全国募集事業費）": ("入学者対策", "協力会補助金（全国募集事業費等）"),
    "協力会補助金（その他）": ("入学者対策", "協力会補助金（全国募集事業費等）"),
}

# 「消防設備保守委託料」「NHK使用料」は寄宿舎系の2事業（鹿砦寮／Penguin Dormitory）で
# 内容文字列が重複するため、内容だけでは判別できない。行番号（45行目以降＝
# Penguin Dormitory側）で振り分ける。鹿砦寮側（学生寮運営）はユーザー指示により
# 集計から除外するため None を返す。
_FINANCE_ROW_TIEBREAK = {
    "消防設備保守委託料": (None, ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）")),
    "NHK使用料": (("通学環境整備", "シェアハウス運営（Penguin House）"), ("通学環境整備", "高校寄宿舎事業（Penguin Dormitory）")),
}


def _finance_classify(content, row_num):
    """「内容」列の文字列を (支援区分, 事業ラベル) に分類する。
    未知の内容は「未分類」区分にフォールバックし、集計から漏れないようにする。
    None を返した場合はユーザー指示による意図的な除外（学生寮運営「鹿砦寮」など）。"""
    content = str(content).strip()
    tiebreak = _FINANCE_ROW_TIEBREAK.get(content)
    if tiebreak:
        return tiebreak[1] if row_num >= 45 else tiebreak[0]
    return FINANCE_CONTENT_CATEGORY.get(content, ("未分類", content))


# 「コーディネーター（宗雲③）」は元データ上は1名分の予算だが、実際は同じ役割で
# 2名分任用しているとのユーザー指示により、金額を2倍して2名分として計上する。
_FINANCE_AMOUNT_MULTIPLIER = {
    "コーディネーター（宗雲③）": 2,
}


def _finance_is_red(cell):
    """フォント色が明示的に赤（元データで高校支援と無関係として着色）かどうか判定する。"""
    color = cell.font.color if cell.font else None
    if color is None or color.type != "rgb" or not color.rgb:
        return False
    return color.rgb.upper().endswith("FF0000")


def aggregate_financial_support():
    finance_path = latest_file(["高校/*高校関連事業一覧*.xlsx"])
    if not finance_path:
        return not_ready()

    try:
        wb = openpyxl.load_workbook(finance_path, data_only=True)
        ws = wb["R8_前年比"]

        # 節別内訳を再掲した重複ピボット表（R7列に"予算額"という文字列が来る行）の
        # 手前までを本体データとして扱う
        end_row = ws.max_row + 1
        for r in range(4, ws.max_row + 1):
            if str(ws.cell(row=r, column=10).value).strip() == "予算額":
                end_row = r
                break

        totals = {}  # (支援区分, 事業ラベル) -> {"r7":.., "r8":..}
        fund_totals = {k: 0 for k in FUND_COLUMNS}
        cur_cat = None
        unclassified = []

        for r in range(4, end_row):
            cat_cell = ws.cell(row=r, column=1).value
            if cat_cell is not None:
                cur_cat = str(cat_cell).strip()
            if cur_cat == "国際バカロレア事業":
                # 鹿追高校支援とは無関係な別事業のため区分ごと除外する
                continue

            content_cell = ws.cell(row=r, column=9)
            content = content_cell.value
            if content is None or str(content).strip() in ("小計", "計", "合計"):
                continue
            if _finance_is_red(content_cell):
                # 元データで赤色着色＝高校支援と無関係として明示された行は除外する
                continue

            classified = _finance_classify(content, r)
            if classified is None:
                # ユーザー指示による意図的な除外（例: 学生寮運営「鹿砦寮」）
                continue
            category, label = classified
            if category == "未分類":
                unclassified.append((r, str(content).strip()))

            multiplier = _FINANCE_AMOUNT_MULTIPLIER.get(str(content).strip(), 1)
            r7 = (ws.cell(row=r, column=10).value or 0) * multiplier
            r8 = (ws.cell(row=r, column=11).value or 0) * multiplier
            bucket = totals.setdefault((category, label), {"r7": 0, "r8": 0})
            bucket["r7"] += int(r7)
            bucket["r8"] += int(r8)

            for k, col in FUND_COLUMNS.items():
                v = ws.cell(row=r, column=col).value
                if v:
                    fund_totals[k] += int(v) * multiplier

        if unclassified:
            print(f"[aggregate_financial_support] 未分類の内容が{len(unclassified)}件あります。"
                  f"FINANCE_CONTENT_CATEGORY を更新してください: {unclassified}")

        order = list(SUPPORT_CATEGORY_ORDER)
        if any(c == "未分類" for c, _ in totals):
            order.append("未分類")

        categories = []
        for cat in order:
            projects = [
                {"label": label, "r7": v["r7"], "r8": v["r8"], "diff": v["r8"] - v["r7"]}
                for (c, label), v in totals.items() if c == cat
            ]
            if not projects:
                continue
            projects.sort(key=lambda p: p["r8"], reverse=True)
            categories.append({
                "label": cat,
                "r7": sum(p["r7"] for p in projects),
                "r8": sum(p["r8"] for p in projects),
                "diff": sum(p["diff"] for p in projects),
                "projects": projects,
            })

        total_r7 = sum(c["r7"] for c in categories)
        total_r8 = sum(c["r8"] for c in categories)
        general_fund = fund_totals["一財"]
        grant_total = total_r8 - general_fund

        return {
            "ready": True,
            "unit": "千円",
            "fiscal_year_current": "R8",
            "fiscal_year_prior": "R7",
            "total_r7": total_r7,
            "total_r8": total_r8,
            "total_diff": total_r8 - total_r7,
            "general_fund": general_fund,
            "grant_total": grant_total,
            "general_fund_ratio_percent": round(general_fund / total_r8 * 100, 1) if total_r8 else None,
            "fund_breakdown": fund_totals,
            "categories": categories,
            "source_updated": datetime.fromtimestamp(os.path.getmtime(finance_path)).strftime("%Y-%m-%d"),
        }
    except Exception as e:
        return not_ready(f"集計エラー: {e}")


# ---------------------------------------------------------------------------
# 5. 入寮者情報
# ---------------------------------------------------------------------------
PREFECTURES = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
    "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
    "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]


def infer_prefecture(address):
    if not isinstance(address, str):
        return "不明"
    addr = unicodedata.normalize("NFKC", address).strip()
    for pref in PREFECTURES:
        if addr.startswith(pref) or pref in addr[:6]:
            return pref
    if re.search(r"[A-Za-z]", addr) and not re.search(r"[都道府県]", addr[:6]):
        return "海外"
    return "不明"


# 施設の定員（生徒名簿からは求められないため固定値。参考: 視察資料側の集計スクリプト）
DORMITORY_CAPACITY = {
    "ペンギンドミトリー": {"male": 20, "female": 20, "total": 40},
    "ペンギンハウス": {"total": 33},  # ハウスは性別ごとの定員なし
}


def aggregate_dormitory():
    result = {}
    dorm_path = latest_file(["高校/*寮*入寮者*.csv"])
    if not dorm_path:
        return {"summary": not_ready()}

    try:
        df = pd.read_csv(dorm_path)
        df.columns = ["facility", "year", "grade", "gender", "address"][: len(df.columns)]
        df["facility"] = df["facility"].astype(str).str.strip().apply(lambda x: unicodedata.normalize("NFKC", x))
        df["grade"] = df["grade"].astype(str).str.strip()
        df["gender"] = df["gender"].astype(str).str.strip()
        df["prefecture"] = df["address"].apply(infer_prefecture)

        grade_order = ["1年生", "2年生", "3年生"]
        facilities = sorted(df["facility"].unique())
        present_prefs = set(df["prefecture"])
        pref_order = [p for p in PREFECTURES if p in present_prefs]
        for extra in ("海外", "不明"):
            if extra in present_prefs:
                pref_order.append(extra)

        by_facility_list = []
        for facility in facilities:
            fdf = df[df["facility"] == facility]
            male = int((fdf["gender"] == "男").sum())
            female = int((fdf["gender"] == "女").sum())
            capacity = DORMITORY_CAPACITY.get(facility, {})
            male_by_grade = [{"grade": g, "count": int(((fdf["gender"] == "男") & (fdf["grade"] == g)).sum())} for g in grade_order]
            female_by_grade = [{"grade": g, "count": int(((fdf["gender"] == "女") & (fdf["grade"] == g)).sum())} for g in grade_order]
            prefecture_matrix = []
            for pref in pref_order:
                pdf = fdf[fdf["prefecture"] == pref]
                counts = {}
                row_total = 0
                for g in grade_order:
                    male_c = int(((pdf["gender"] == "男") & (pdf["grade"] == g)).sum())
                    female_c = int(((pdf["gender"] == "女") & (pdf["grade"] == g)).sum())
                    counts[g] = {"male": male_c, "female": female_c}
                    row_total += male_c + female_c
                prefecture_matrix.append({"prefecture": pref, "counts": counts, "total": row_total})
            by_facility_list.append({
                "name": facility,
                "count": male + female,
                "male": male,
                "female": female,
                "capacity_male": capacity.get("male"),
                "capacity_female": capacity.get("female"),
                "capacity_total": capacity.get("total"),
                "male_by_grade": male_by_grade,
                "female_by_grade": female_by_grade,
                "prefecture_matrix": prefecture_matrix,
            })
        by_facility_list.sort(key=lambda x: -x["count"])

        by_grade_list = [
            {
                "grade": g,
                "count": int((df["grade"] == g).sum()),
                "male": int(((df["grade"] == g) & (df["gender"] == "男")).sum()),
                "female": int(((df["grade"] == g) & (df["gender"] == "女")).sum()),
            }
            for g in grade_order
        ]
        by_prefecture_list = [
            {
                "prefecture": pref,
                "count": int((df["prefecture"] == pref).sum()),
                "male": int(((df["prefecture"] == pref) & (df["gender"] == "男")).sum()),
                "female": int(((df["prefecture"] == pref) & (df["gender"] == "女")).sum()),
            }
            for pref in sorted(present_prefs, key=lambda p: -int((df["prefecture"] == p).sum()))
        ]

        total_male = sum(f["male"] for f in by_facility_list)
        total_female = sum(f["female"] for f in by_facility_list)

        result["summary"] = {
            "ready": True,
            "total": int(len(df)),
            "male": total_male,
            "female": total_female,
            "by_facility": by_facility_list,
            "by_grade": by_grade_list,
            "by_prefecture": by_prefecture_list,
            "prefecture_order": pref_order,
            "grade_order": grade_order,
            "source_updated": datetime.fromtimestamp(os.path.getmtime(dorm_path)).strftime("%Y-%m-%d"),
        }
    except Exception as e:
        result["summary"] = not_ready(f"集計エラー: {e}")

    return result


# ---------------------------------------------------------------------------
# メイン処理
# ---------------------------------------------------------------------------
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    career_result, charm_result = aggregate_career_and_learning()

    data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "school_overview": aggregate_school_overview(),
        "education_activities": aggregate_education_activities(charm_result),
        "career_and_learning": career_result,
        "financial_support": aggregate_financial_support(),
        "dormitory": aggregate_dormitory(),
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # index.html を file:// で直接開いても読み込めるよう、
    # fetch()を使わずに済むJS埋め込み版も同時に出力する。
    with open(OUTPUT_JS_FILE, "w", encoding="utf-8") as f:
        f.write("window.DASHBOARD_DATA = ")
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    print(f"生成しました: {OUTPUT_FILE}")
    print(f"生成しました: {OUTPUT_JS_FILE}")


if __name__ == "__main__":
    main()
